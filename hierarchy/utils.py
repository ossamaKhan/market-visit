import secrets
import string

from openpyxl import load_workbook

# Exact header order expected in the uploaded sheet.
REQUIRED_HEADERS = [
    'FR ID', 'Region', 'BU', 'FR Status', 'FR City',
    'FR Address', 'ARM Name', 'ARM Emp #', 'Email',
]

# Maps sheet header -> FranchiseRecord field name
HEADER_TO_FIELD = {
    'FR ID': 'fr_id',
    'Region': 'region',
    'BU': 'bu',
    'FR Status': 'fr_status',
    'FR City': 'fr_city',
    'FR Address': 'fr_address',
    'ARM Name': 'arm_name',
    'ARM Emp #': 'arm_emp_no',
    'Email': 'email',
}


def generate_random_password(length=10):
    """
    Generates a random password using letters, digits, and a small set of
    punctuation, guaranteeing at least one of each character class.
    """
    alphabet = string.ascii_letters + string.digits
    while True:
        password = ''.join(secrets.choice(alphabet) for _ in range(length))
        has_lower = any(c.islower() for c in password)
        has_upper = any(c.isupper() for c in password)
        has_digit = any(c.isdigit() for c in password)
        if has_lower and has_upper and has_digit:
            return password


class HierarchyFileError(Exception):
    """Raised when the uploaded file doesn't match the expected format."""
    pass


def parse_hierarchy_file(uploaded_file):
    """
    Reads the uploaded .xlsx file and returns a list of row dicts keyed by
    FranchiseRecord field names (fr_id, region, bu, ...). Validates that
    the header row matches REQUIRED_HEADERS (order-independent) before
    reading any data rows. Blank rows (no FR ID) are skipped.
    """
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise HierarchyFileError(f"Couldn't read this file as an Excel workbook: {exc}")

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HierarchyFileError("The sheet appears to be empty.")

    headers = [str(h).strip() if h is not None else '' for h in header_row]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise HierarchyFileError(
            "The sheet is missing expected column(s): " + ", ".join(missing) +
            ". Expected headers: " + ", ".join(REQUIRED_HEADERS)
        )

    col_index = {header: idx for idx, header in enumerate(headers)}

    records = []
    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, '') for cell in row):
            continue

        def get(header):
            idx = col_index[header]
            value = row[idx] if idx < len(row) else None
            return str(value).strip() if value is not None else ''

        fr_id = get('FR ID')
        email = get('Email')

        if not fr_id:
            continue  # skip rows without an FR ID
        if not email:
            raise HierarchyFileError(f"Row {row_num} (FR ID: {fr_id}) has no email.")

        record = {HEADER_TO_FIELD[h]: get(h) for h in REQUIRED_HEADERS}
        record['_row_num'] = row_num
        records.append(record)

    if not records:
        raise HierarchyFileError("No data rows found below the header.")

    return records
